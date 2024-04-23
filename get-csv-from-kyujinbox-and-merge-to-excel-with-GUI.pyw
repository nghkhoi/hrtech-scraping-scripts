import wx
import wx.grid
import os
import csv
import openpyxl
import codecs
import subprocess
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

USERNAME = ''
PASSWORD = ''

LOGIN_URL = ''

CONFIG_FILE_PATH = './config.ini'

class MyFrame(wx.Frame):
    def __init__(self):
        super().__init__(None, title="KYUJINBOXヘルパー", size=(700, 500))
        
        panel = wx.Panel(self)

        self.text_edit = wx.TextCtrl(panel, style=wx.TE_MULTILINE)
        
        font = self.text_edit.GetFont()
        font.SetPointSize(13)  # Set font size to 12
        font.SetWeight(wx.FONTWEIGHT_BOLD)  # Set bold
        self.text_edit.SetFont(font)

        self.selected_folder_path, text_value = self.load_configuration()
        self.text_edit.SetValue(text_value)

        self.folder_picker = wx.DirPickerCtrl(panel, message="フォルダーを選んでください。", path=self.selected_folder_path, style=wx.DIRP_DEFAULT_STYLE)
        picker_btn = self.folder_picker.GetPickerCtrl()
        if picker_btn is not None:
            picker_btn.SetLabel("参照")
        self.folder_picker.Bind(wx.EVT_DIRPICKER_CHANGED, self.on_folder_pick)

        button = wx.Button(panel, label="実行")
        font = button.GetFont()
        font.PointSize = 14
        button.SetFont(font)
        button.SetBackgroundColour((19, 27, 63, 255))
        button.SetForegroundColour((255, 255, 255, 255))
        button.Bind(wx.EVT_BUTTON, self.proceed)

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.text_edit, proportion=1, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(self.folder_picker, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(button, flag=wx.ALIGN_CENTER | wx.ALL, border=10)
        
        panel.SetSizer(sizer)

        self.text_edit.SetFocus()
        self.text_edit.SelectAll()
        
    def load_configuration(self):
        selected_folder_path = wx.StandardPaths.Get().GetDocumentsDir()
        text_value = ""
        
        if os.path.exists(CONFIG_FILE_PATH):
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as config_file:
                selected_folder_path = config_file.readline().strip()
                text_value = config_file.read()

        return selected_folder_path, text_value

    def save_configuration(self):
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as config_file:
            config_file.write(self.selected_folder_path + '\n')
            config_file.write(self.text_edit.GetValue())
        
    def on_folder_pick(self, event):
        self.selected_folder_path = self.folder_picker.GetPath()

    def proceed(self, event):
        self.text_edit.SetDefaultStyle(wx.TextAttr(wx.NullColour, wx.NullColour))

        text_data = self.text_edit.GetValue()
        self.save_configuration()

        lines = text_data.split('\n')

        # Windowsセレニウム起動

        service = Service(executable_path=r'C:\chromedriver\chromedriver.exe')
        chrome_options = Options()
        chrome_options.add_argument('--lang=ja-JP')
        chrome_options.add_experimental_option('detach', True)
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": self.selected_folder_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        })
        driver = webdriver.Chrome(service=service, options=chrome_options)

        # Windowsセレニウム起動

        driver.get(LOGIN_URL)
        sleep(2)

        try:
            el_username = driver.find_element(By.CSS_SELECTOR, "input#login_email")
            el_username.send_keys(USERNAME)
            sleep(2)

            el_password = driver.find_element(By.CSS_SELECTOR, "input#login_password")
            el_password.send_keys(PASSWORD)
            sleep(2)

            #wx.MessageBox("This is a simple alert!", "Alert", wx.OK | wx.ICON_INFORMATION)

            btn_submit = driver.find_element(By.CSS_SELECTOR, "button#BtnLogin")
            btn_submit.submit()
            sleep(1)

            #print('ユーザー名、パスワードでログイン済')

            driver.maximize_window()

            for line in lines:

                variables = line.split(',')

                if len(variables) != 5:
                    continue

                driver.get('https://saiyo.kyujinbox.com/ptr/l-accounts')

                company_name, confirm_type, category, start_date, end_date = map(str.strip, variables)

                el_confirm = driver.find_element(By.XPATH, f"//a[contains(@class, 'c-switch__item') and contains(text(), '{confirm_type}')]")
                el_confirm.click();
                
                sleep(1);

                el_company = driver.find_element(By.XPATH, f"//a[contains(@class, 'c-link') and contains(text(), '{company_name}')]")
                el_company.click();

                if confirm_type == '直接投稿':
                    el_adredirect = driver.find_element(By.XPATH, "//a[contains(@class, 'Ad') and contains(text(), '有料利用状況')]")
                    el_adredirect.click();
                
                    el_reporting = driver.find_element(By.XPATH, "//a[contains(@class, 'c-tab__link') and contains(text(), 'レポーティング')]")
                    el_reporting.click();
                
                    el_category = driver.find_element(By.XPATH, f"//p[contains(@class, 'TabBtn') and contains(text(), '{category}')]")
                    el_category.click();
                
                    el_dateselect = driver.find_element(By.CSS_SELECTOR, ".ReportFilterForm > .DateModal-select")
                    el_dateselect.click();

                    ip_from = driver.find_element(By.CSS_SELECTOR, "input[name='modal_from']")
                    sleep(1)
                    ip_from.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
                    ip_from.send_keys(start_date)
                    ip_from.send_keys(Keys.TAB)
                    ip_to = driver.find_element(By.CSS_SELECTOR, "input[name='modal_to']")
                    sleep(1)
                    ip_to.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
                    ip_to.send_keys(end_date)
                    ip_from.send_keys(Keys.TAB)
                    btn_submit = driver.find_element(By.CSS_SELECTOR, ".DateModal_footer > button.BtnBlue")
                    sleep(2)
                    btn_submit.click()

                elif confirm_type == 'クローリング・フィード':
                    el_report = driver.find_element(By.XPATH, "//li/a[contains(text(), 'レポート')]")
                    el_report.click();
                
                    el_category = driver.find_element(By.XPATH, f"//ul/li[contains(text(), '{category}')]")
                    el_category.click();
                
                    el_dateselect = driver.find_element(By.CSS_SELECTOR, ".displayPeriod > .modal-date-select")
                    el_dateselect.click();

                    ip_from = driver.find_element(By.CSS_SELECTOR, "input[name='modal_from']")
                    sleep(1)
                    ip_from.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
                    ip_from.send_keys(start_date)
                    ip_from.send_keys(Keys.TAB)
                    ip_to = driver.find_element(By.CSS_SELECTOR, "input[name='modal_to']")
                    sleep(1)
                    ip_to.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
                    ip_to.send_keys(end_date)
                    ip_to.send_keys(Keys.TAB)
                    btn_submit = driver.find_element(By.CSS_SELECTOR, "button.js-apply-btn")
                    sleep(2)
                    btn_submit.click()
                
                sleep(1)

                el_download = driver.find_element(By.CSS_SELECTOR, "a.js-csv-download")
                el_download.click()

            wx.MessageBox(f'ステップ（1/2）　会社CSVファイル　ダウンロード完了！', 'メッセージー', wx.OK | wx.ICON_INFORMATION)

            merged_filename = os.path.join(self.selected_folder_path, "MERGED.xlsx")

            if os.path.exists(merged_filename):
                message_dialog = wx.MessageDialog(self, "MERGED.xlsx が既に存在します。上書きしますか？", "警告", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_WARNING)
                if message_dialog.ShowModal() != wx.ID_YES:
                    message_dialog.Destroy()
                    return
                message_dialog.Destroy()

            merged_workbook = openpyxl.Workbook()
            default_sheet = merged_workbook.active
            default_sheet.title = "デフォルト"

            for filename in os.listdir(self.selected_folder_path):
                if filename.endswith(".csv"):
                    csv_path = os.path.join(self.selected_folder_path, filename)
                    sheet_title = filename.split('_')[0][:31] if '_' in filename else filename[:-4][:31]
                    worksheet = merged_workbook.create_sheet(title=sheet_title)
                    with codecs.open(csv_path, 'r', 'shift_jis') as file:
                        csv_reader = csv.reader(file)
                        for row_index, row in enumerate(csv_reader, start=1):
                            for col_index, value in enumerate(row, start=1):
                                worksheet.cell(row=row_index, column=col_index).value = value.encode('utf-8').decode('utf-8')

            merged_workbook.remove(default_sheet)

            merged_workbook.save(merged_filename)
            wx.MessageBox("ステップ（2/2）　CSVファイルをマージし、MERGED.xlsxとして保存されました。", "成功", wx.OK | wx.ICON_INFORMATION)
            subprocess.Popen(["start", merged_filename], shell=True)

        except Exception as e:
            wx.MessageBox(f'{type(e).__name__} \n {str(e)}', '例外エラー', wx.OK | wx.ICON_ASTERISK)

app = wx.App()
frame = MyFrame()
frame.Show()
app.MainLoop()