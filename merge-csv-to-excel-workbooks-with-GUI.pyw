import wx
import os
import csv
import openpyxl
import codecs
import subprocess

class MyFrame(wx.Frame):
    def __init__(self, parent, title):
        super(MyFrame, self).__init__(parent, title=title, size=(400, 100))

        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        hbox = wx.BoxSizer(wx.HORIZONTAL)

        self.folder_path = wx.TextCtrl(panel, style=wx.TE_READONLY)
        hbox.Add(self.folder_path, proportion=1, flag=wx.EXPAND | wx.ALL, border=10)

        pick_process_button = wx.Button(panel, label="フォルダを選択して処理")
        pick_process_button.Bind(wx.EVT_BUTTON, self.on_pick_process)
        hbox.Add(pick_process_button, flag=wx.EXPAND | wx.ALL, border=10)

        vbox.Add(hbox, proportion=1, flag=wx.EXPAND)

        panel.SetSizer(vbox)

    def on_pick_process(self, event):
        dlg = wx.DirDialog(self, "フォルダを選択してください:", style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST)
        if dlg.ShowModal() == wx.ID_OK:
            folder_path = dlg.GetPath()
            self.folder_path.SetValue(folder_path)
            self.process_csv_files(folder_path)
        dlg.Destroy()

    def process_csv_files(self, folder_path):
        if not os.path.isdir(folder_path):
            wx.MessageBox("無効なフォルダパスです！", "エラー", wx.OK | wx.ICON_ERROR)
            return

        merged_filename = os.path.join(folder_path, "MERGED.xlsx")

        if os.path.exists(merged_filename):
            message_dialog = wx.MessageDialog(self, "MERGED.xlsx が既に存在します。上書きしますか？", "警告", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_WARNING)
            if message_dialog.ShowModal() != wx.ID_YES:
                message_dialog.Destroy()
                return
            message_dialog.Destroy()

        merged_workbook = openpyxl.Workbook()
        default_sheet = merged_workbook.active
        default_sheet.title = "デフォルト"

        for filename in os.listdir(folder_path):
            if filename.endswith(".csv"):
                csv_path = os.path.join(folder_path, filename)
                # Truncate the title if it exceeds 31 characters
                sheet_title = filename[:-4][:31]  
                worksheet = merged_workbook.create_sheet(title=sheet_title)
                with codecs.open(csv_path, 'r', 'shift_jis') as file:
                    csv_reader = csv.reader(file)
                    for row_index, row in enumerate(csv_reader, start=1):
                        for col_index, value in enumerate(row, start=1):
                            # Convert value to UTF-8 before writing to Excel
                            worksheet.cell(row=row_index, column=col_index).value = value.encode('utf-8').decode('utf-8')

        merged_workbook.remove(default_sheet)

        merged_workbook.save(merged_filename)
        wx.MessageBox("CSVファイルをマージし、MERGED.xlsxとして保存されました。", "成功", wx.OK | wx.ICON_INFORMATION)
        subprocess.Popen(["start", merged_filename], shell=True)


class MyApp(wx.App):
    def OnInit(self):
        frame = MyFrame(None, title="CSVマージ")
        frame.Show()
        return True


if __name__ == "__main__":
    app = MyApp(False)
    app.MainLoop()
